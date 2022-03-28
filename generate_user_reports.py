import openpyxl
import os
import populate_database
from config import configured_reports, generate_reports_output_dir, Pesquisador, reports_as_new_worksheets
from configured_reports.write_item_info import write_item_info


def report_is_valid(report):
    entities_in_report = []

    for item in configured_reports[report]:

        try:
            entity = item.split(".")[0]
        except:
            print("The report \"" + report + "\" is not a valid report because " + str(item) + " is not a valid input\n")
            return False

        if item not in Pesquisador.__dict__.values() and entity not in entities_in_report:
            entities_in_report.append(entity)

    if len(entities_in_report) > 1:
        print("The report \"" + report + "\" is not a valid report because it has two or more entities other than \"Pesquisador\"")
        print("These entities are: \"" + ", ".join(entities_in_report) + "\"\n")
        return False

    return True


def make_researcher_cartesian_product(item, report):
    if item in Pesquisador.__dict__.values():
        for other_item in configured_reports[report]:
            if other_item not in Pesquisador.__dict__.values():
                return True, other_item

    return False, None


def write_report_items(col, report, session, worksheet):
    researcher_cartesian_product = (False, None)
    for item in configured_reports[report]:

        if not researcher_cartesian_product[0]: researcher_cartesian_product = make_researcher_cartesian_product(item, report)
        write_item_info(session, item, worksheet, col, researcher_cartesian_product)
        col += 1


def write_reports(session):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for report in configured_reports:
        if report_is_valid(report):

            if not reports_as_new_worksheets: wb = openpyxl.Workbook()

            worksheet = wb.create_sheet(report) if reports_as_new_worksheets else wb.active

            col = 1

            write_report_items(col, report, session, worksheet)

            if not reports_as_new_worksheets: save_report(report, wb)

    if reports_as_new_worksheets: save_report("configured_reports", wb)


def save_report(report, workbook):
    if len(workbook.sheetnames) < 1:
        print("There isn't any report to generate\n")
        return None

    workbook.save(generate_reports_output_dir + os.sep + report + ".xlsx")
    print("Finished generating the report " + report + "\n")


def main():
    session = populate_database.main()

    print("\nGenerating reports...\n")

    write_reports(session)


if __name__ == "__main__":
    main()