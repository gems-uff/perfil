import openpyxl
import os
from database.entities.researcher import Affiliation
import populate_database
from collections import defaultdict
from config import generate_datacapes_output_dir, QualisLevel, datacapes_minimum_similarity_titles
from database.database_manager import Researcher, Journal, Conference, JournalPaper, Venue
from utils.xlsx_utils import calculate_number_of_pages, get_qualis_points
from utils.similarity_manager import get_similarity
from utils.list_filters import scope_years_paper_or_support, published_journal_paper, jcr_pub_filter, \
    qualis_level_journal, qualis_level_conference, conference_paper, journal_paper, affiliated_researcher, completed_paper_filter


def filter_completed_scope_years_papers(paper_list):
    return list(filter(completed_paper_filter, list(filter(scope_years_paper_or_support, paper_list))))


def write(worksheet, row, column, value):
    worksheet.cell(row, column, value)
    return column + 1


def write_production_header(worksheet, write_researcher : bool):
    """Writes the header of the xlsx file"""
    if(write_researcher):
        worksheet.cell(row=1, column=1, value="Docente")
        worksheet.cell(row=1, column=2, value="Ultima atualizacao Lattes")
    offset = (2 if write_researcher else 0)
    worksheet.cell(row=1, column=1+offset, value="Ano")
    worksheet.cell(row=1, column=2+offset, value="Tipo")
    worksheet.cell(row=1, column=3+offset, value="Titulo")
    worksheet.cell(row=1, column=4+offset, value="Autores")
    worksheet.cell(row=1, column=5+offset, value="Paginas")
    worksheet.cell(row=1, column=6+offset, value="DOI")
    worksheet.cell(row=1, column=7+offset, value="Venue")
    worksheet.cell(row=1, column=8+offset, value="Match")
    worksheet.cell(row=1, column=9+offset, value="ISSN/Sigla Evento")
    worksheet.cell(row=1, column=10+offset, value="JCR")
    worksheet.cell(row=1, column=11+offset, value="Qualis")
    worksheet.cell(row=1, column=12+offset, value="Pontos")


def write_production_paper(paper, researcher, row, venue, worksheet, is_journal_paper : bool, write_researcher : bool):
    """Writes info about papers in a xlsx file"""
    if(write_researcher):
        worksheet.cell(row=row, column=1, value=researcher.name)
        worksheet.cell(row=row, column=2, value=researcher.last_lattes_update)
    offset = (2 if write_researcher else 0)
    worksheet.cell(row=row, column=1+offset, value=paper.year)
    worksheet.cell(row=row, column=2+offset, value="PERIODICO" if is_journal_paper else "CONFERENCIA")
    worksheet.cell(row=row, column=3+offset, value=paper.title)
    worksheet.cell(row=row, column=4+offset, value=paper.authors)
    worksheet.cell(row=row, column=5+offset, value=calculate_number_of_pages(paper))
    worksheet.cell(row=row, column=6+offset, value=" https://doi.org/" + paper.doi if paper.doi is not None and paper.doi.strip() != "" else "null")
    worksheet.cell(row=row, column=7+offset, value=venue.name)
    worksheet.cell(row=row, column=8+offset, value=venue.official_forum)  
    worksheet.cell(row=row, column=9+offset, value=venue.issn if is_journal_paper else venue.acronym)
    worksheet.cell(row=row, column=10+offset, value=venue.jcr if is_journal_paper else "N/A")
    worksheet.cell(row=row, column=11+offset, value=venue.qualis.value)
    worksheet.cell(row=row, column=12+offset, value=get_qualis_points(is_journal_paper, venue.qualis))


def researcher_production_paper_iterator(array_papers, researcher, session, worksheet, row, is_journal_paper : bool):
    """For each paper of a given researcher calls the function to write its info"""
    for i in range(len(array_papers)):
        paper = array_papers[i]
        venue = paper.venue

        write_production_paper(paper, researcher, row, venue, worksheet, is_journal_paper, True)
        row += 1  # next paper info must be written on next row

    return row  # returns the last roll after all the papers' info


def write_researchers_production(researchers, session):
    """Writes the producao_docentes.xlsx file"""
    wb = openpyxl.Workbook()
    worksheet = wb.active
    write_production_header(worksheet, True)
    row = 2

    papers = []  # array to get and filter all the papers only once
    for researcher in researchers:
        conference_papers = filter_completed_scope_years_papers(researcher.conference_papers)
        row = researcher_production_paper_iterator(conference_papers, researcher, session, worksheet, row, False)

        journal_papers = list(filter(published_journal_paper, filter_completed_scope_years_papers(researcher.journal_papers)))
        row = researcher_production_paper_iterator(journal_papers, researcher, session, worksheet, row, True)

        papers += list(filter(lambda x: affiliated_researcher(researcher.id, x.year, session), conference_papers))  # Adds the conference papers of the researcher to the list
        papers += list(filter(lambda x: affiliated_researcher(researcher.id, x.year, session), journal_papers))  # Adds the journal papers of the researcher to the list

    wb.save(generate_datacapes_output_dir + os.sep + "producao_docentes.xlsx")

    return papers  # Return all the papers of the researchers in the scope years


def write_yearly_production(papers, session):
    """Writes the producao_anual.xlsx file"""
    papers.sort(key=lambda paper: (paper.year, paper.type))

    wb = openpyxl.Workbook()
    worksheet = wb.active
    write_production_header(worksheet, False)

    for i in range(len(papers)):
        is_journal_paper = isinstance(papers[i], JournalPaper)
        venue = papers[i].venue

        write_production_paper(papers[i], None, i + 2, venue, worksheet, is_journal_paper, False)

    wb.save(generate_datacapes_output_dir + os.sep + "producao_anual.xlsx")


def write_summary_header_papers(column, periodico, worksheet):
    for qualis_level in QualisLevel:
        worksheet.cell(row=1, column=column, value=periodico + " " + qualis_level.value)
        column += 1
    return column


def write_paper_number_by_qualis(session, column, papers, row, worksheet, is_journal_list : bool):
    """Writes the amount of papers of each qualis in a .xlsx file.
    It also calculates the restricted index, general index and the index of (a1, a2, b1, b2) qualis"""
    restricted_index = 0
    general_index = 0
    filter_function = qualis_level_journal if is_journal_list else qualis_level_conference

    for qualis_level in QualisLevel:
        papers_with_this_qualis_level = list(filter(lambda x: filter_function(x, session, qualis_level), papers))
        worksheet.cell(row=row, column=column, value=len(papers_with_this_qualis_level))

        # Calculates restricted_index and general_index
        qualis_points = get_qualis_points(is_journal_list, qualis_level)
        if qualis_level in [QualisLevel.A1, QualisLevel.A2, QualisLevel.A3, QualisLevel.A4]: restricted_index += len(papers_with_this_qualis_level) * qualis_points
        general_index += len(papers_with_this_qualis_level) * qualis_points

        column += 1

    return restricted_index, general_index


def write_summary_header(worksheet, entity, per_researcher=False):
    """Writes the header of datacapes summary files"""
    column = write(worksheet, 1, 1, entity)
    if per_researcher:
        column = write(worksheet, 1, column, 'DOCENTES')
    column = write(worksheet, 1, column, 'PERIODICOS JCR')    
    column = write_summary_header_papers(column, "PERIODICO", worksheet)
    column = write_summary_header_papers(column, "CONFERENCIA", worksheet)
    column = write(worksheet, 1, column, 'I-RESTRITO PERIODICO')
    column = write(worksheet, 1, column, 'I-RESTRITO CONFERENCIA')
    column = write(worksheet, 1, column, 'I-RESTRITO TOTAL')
    if per_researcher:
        column = write(worksheet, 1, column, 'I-RESTRITO POR DOCENTE')
    column = write(worksheet, 1, column, 'I-GERAL PERIODICO')
    column = write(worksheet, 1, column, 'I-GERAL CONFERENCIA')
    column = write(worksheet, 1, column, 'I-GERAL TOTAL')
    if per_researcher:
        column = write(worksheet, 1, column, 'I-GERAL POR DOCENTE')
    column = write(worksheet, 1, column, 'SATURACAO (TRAVA)')


def write_summary(conference_papers, journal_papers, journal_papers_jcr, entity, per_researcher, row, session, worksheet):
    """Writes the quantitative info about papers in a .xlsx file"""
    column = write(worksheet, row, 1, entity)
    if per_researcher:
        num_researchers = session.query(Affiliation).filter(Affiliation.year == entity).count()
        column = write(worksheet, row, column, num_researchers)
    column = write(worksheet, row, column, len(journal_papers_jcr))
    journal_indexes = write_paper_number_by_qualis(session, column, journal_papers, row, worksheet, True)
    conference_indexes = write_paper_number_by_qualis(session, len(QualisLevel) + column, conference_papers, row, worksheet, False)

    # get indexes from tuple
    journal_restricted_index = journal_indexes[0]
    conference_restricted_index = conference_indexes[0]
    restricted_index = journal_restricted_index + conference_restricted_index
    journal_general_index = journal_indexes[1]
    conference_general_index = conference_indexes[1]
    general_index = journal_general_index + conference_general_index

    column = len(QualisLevel) * 2 + column # column number jump after writing the number of journals and conferences by qualis

    column = write(worksheet, row, column, journal_restricted_index)
    column = write(worksheet, row, column, conference_restricted_index)
    column = write(worksheet, row, column, restricted_index)
    if per_researcher:
        column = write(worksheet, row, column, restricted_index / num_researchers)
    column = write(worksheet, row, column, journal_general_index)
    column = write(worksheet, row, column, conference_general_index)
    column = write(worksheet, row, column, general_index)
    if per_researcher:
        column = write(worksheet, row, column, general_index / num_researchers)

    # SATURACAO (TRAVA)
    if journal_general_index != 0:
        column = write(worksheet, row, column, conference_general_index / journal_general_index)
    else:
        if conference_general_index != 0:
            column = write(worksheet, row, column, "INF.")
        else:
            column = write(worksheet, row, column, "IND.")


def write_researchers_summary(researchers, session):
    """Writes the sumario_docentes.xlsx file"""
    wb = openpyxl.Workbook()
    worksheet = wb.active
    write_summary_header(worksheet, 'DOCENTE', False)
    row = 2  # the header uses the first row
    for researcher in researchers:
        conference_papers = filter_completed_scope_years_papers(researcher.conference_papers)
        journal_papers = list(filter(published_journal_paper, filter_completed_scope_years_papers(researcher.journal_papers)))
        journal_papers_jcr = list(filter(lambda x: jcr_pub_filter(x), journal_papers))

        write_summary(conference_papers, journal_papers, journal_papers_jcr, researcher.name, False, row, session, worksheet)

        row += 1  # a row for each researcher
    wb.save(generate_datacapes_output_dir + os.sep + "sumario_docentes.xlsx")


def write_yearly_summary(papers, session):
    """Writes the sumario_anual.xslx file"""

    # Sort papers by year to make sure
    papers_by_year = defaultdict(list)
    for paper in papers:
        papers_by_year[paper.year].append(paper)

    wb = openpyxl.Workbook()
    worksheet = wb.active
    write_summary_header(worksheet, 'ANO', True)
    row = 2

    for year in papers_by_year:
        conference_papers = list(filter(conference_paper, papers_by_year[year]))
        journal_papers = list(filter(journal_paper, papers_by_year[year]))
        journal_papers_jcr = list(filter(lambda x: jcr_pub_filter(x), journal_papers))

        write_summary(conference_papers, journal_papers, journal_papers_jcr, year, True, row, session, worksheet)

        row += 1 # a row for each year
    wb.save(generate_datacapes_output_dir + os.sep + "sumario_anual.xlsx")


def remove_paper_duplicates(papers_list, session):
    """Removes duplicates, same title, of the same papers"""
    papers = list(set(papers_list))
    new_list = papers.copy()

    for i in range(len(papers)):
        if papers[i] in new_list:
            venue_i = papers[i].venue
            for j in range(i + 1, len(papers)):
                venue_j = papers[j].venue
                same_venue = venue_i.official_forum == venue_j.official_forum if (venue_i.official_forum is not None) and (venue_j.official_forum is not None) else venue_i.name.lower() == venue_j.name.lower()
                same_year = papers[i].year == papers[j].year

                if same_venue and same_year and (papers[j] in new_list) \
                        and (papers[i].title.lower() == papers[j].title.lower() or get_similarity(papers[i].title.lower(), papers[j].title.lower()) >= datacapes_minimum_similarity_titles):
                    if isinstance(papers[i], JournalPaper) and isinstance(papers[j], JournalPaper) and (venue_i.jcr is not None) and (venue_j.jcr is not None) and (venue_i.jcr < venue_j.jcr) and (papers[i] in new_list):
                        new_list.remove(papers[i])  # removes the paper with the wrong issn
                    else:
                        new_list.remove(papers[j])  # if there are two papers with the same title, two researchers worked on the same paper, one gets remove

    return new_list  # the list with only a paper of each


def main():
    session = populate_database.main()
    researchers = session.query(Researcher).all()

    print("\nGenerating datacapes files...\n")

    # writes researchers production
    papers = remove_paper_duplicates(write_researchers_production(researchers, session), session)  # Some files only need a paper of each, they don't make distiction between the researchers

    print("producao_docentes.xlsx was generated\n")
    # writes yearly production
    write_yearly_production(papers, session)  # here it sorts the papers by year which is used on the next files
    print("producao_anual.xlsx was generated\n")
    # writes researcher summary
    write_researchers_summary(researchers, session)
    print("sumario_docentes.xlsx was generated\n")
    # writes yearly summary
    write_yearly_summary(papers, session)
    print("sumario_anual.xlsx was generated\n")

    print("Finished generating the datacapes files\n")


if __name__ == "__main__":
    main()