import openpyxl
import os
import populate_database
from config import configured_reports, generate_reports_output_dir, Pesquisador, reports_as_new_worksheets, new_worksheet_if_conflict
from configured_reports.write_item_info import write_item_info


def get_entities_in_report(report):
    """Returns all the entities in a report"""
    entities_in_report = []

    for item in configured_reports[report]:

        try:
            entity = item.split(".")[0]
        except:
            return []

        if item not in Pesquisador.__dict__.values() and entity not in entities_in_report:
            entities_in_report.append(entity)

    return entities_in_report


def report_is_valid(entities_in_report, report):
    """Checks if a report written by the user is valid"""

    if len(entities_in_report) == 0:
        print("The report \"" + report + "\" is not a valid report because there is an unacceptable input in it\n")
        return False

    if len(entities_in_report) > 1 :
        if (not new_worksheet_if_conflict) or reports_as_new_worksheets:
            print("The report \"" + report + "\" is not a valid report because it has two or more entities other than \"Pesquisador\"")
            print("These entities are: \"" + ", ".join(entities_in_report) + "\"\n")
        return False

    return True


def invalid_report_to_worksheets(entities_in_report, report):
    """Returns all the entities and their attributes, other than Pesquisador, in a invalid report as a dict's keys to be written in the
    same workbook as different worksheets"""

    entities_as_worksheets = dict()

    for entity in entities_in_report:
        entities_as_worksheets[entity] = []
        researcher_attribute_index = 0

        for item in configured_reports[report]:
            if entity in item:
                entities_as_worksheets[entity].append(item)
            elif item in Pesquisador.__dict__.values():
                entities_as_worksheets[entity].insert(researcher_attribute_index, item)
                researcher_attribute_index += 1

    return entities_as_worksheets


def make_researcher_cartesian_product(item, report_list):
    """Checks if an attribute of the researcher class/entity needs to be written as a cartesian production with
    another class/entity and returns a tuple in the format (bool, user_class.attribute)"""

    if item in Pesquisador.__dict__.values():
        for other_item in report_list:
            if other_item not in Pesquisador.__dict__.values():
                return True, other_item

    return False, None


def write_report_items(col, report_list, session, worksheet):
    """For each item in a report calls the function to write it in a specific column"""

    researcher_cartesian_product = (False, None)
    for item in report_list:

        if not researcher_cartesian_product[0]: researcher_cartesian_product = make_researcher_cartesian_product(item, report_list)
        write_item_info(session, item, worksheet, col, researcher_cartesian_product)
        col += 1


def write_a_report(report_name, report_list, new_sheet, session, wb):
    """Writes only one report"""

    worksheet = wb.create_sheet(report_name) if new_sheet else wb.active

    col = 1

    write_report_items(col, report_list, session, worksheet)

    return wb


def write_reports(session):
    """Gets the reports from the config.py file and for each report calls the function to write its items,
    be it as a new file or as a new sheet. Then calls the function to save the report(s)"""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for report in configured_reports:
        entities_in_report = get_entities_in_report(report)
        valid_report = report_is_valid(entities_in_report, report)

        if not reports_as_new_worksheets: wb = openpyxl.Workbook()

        if (not valid_report) and (not reports_as_new_worksheets) and new_worksheet_if_conflict and len(entities_in_report) > 0:
            report_to_worksheets = invalid_report_to_worksheets(entities_in_report, report)

            wb.remove(wb.active)

            for entity_report in report_to_worksheets:
                write_a_report(entity_report, report_to_worksheets[entity_report], True, session, wb)

            save_report(report, wb)

        elif valid_report:

            wb = write_a_report(report, configured_reports[report], reports_as_new_worksheets, session, wb)
            if not reports_as_new_worksheets: save_report(report, wb)

    if reports_as_new_worksheets: save_report("configured_reports", wb)


def save_report(report, workbook):
    """Checks if there is a report to generate and save it as a .xslx file"""

    if len(workbook.sheetnames) < 1:
        print("There isn't any report to generate\n")
        return None

    workbook.save(generate_reports_output_dir + os.sep + report + ".xlsx")
    print("Finished generating the report " + report + "\n")


def main():
    session = populate_database.main()

    print("\nGenerating reports...\n")

    write_reports(session)


if __name__ == "__main__":
    main()