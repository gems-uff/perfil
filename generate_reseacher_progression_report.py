import openpyxl
import os
import populate_database
import argparse
import pandas as pd
from config import QualisLevel, generate_reseacher_progression_report_output_dir, researchers_file
from database.database_manager import Researcher, Journal, Conference, Advisement, Committee
from database.entities.titles_support import CommitteeTypes
from utils.list_filters import scope_years_paper_or_support
from utils.xlsx_utils import calculate_number_of_pages, get_qualis_points


def researchers_from_args(researcher, arg_id_list):
    """Filter function to get the researchers ids or their  lattes_ids from the arguments"""
    return (researcher.id in arg_id_list) or (researcher.lattes_id in arg_id_list)


def write_journal_papers(researcher, session, workbook):
    """Writes a sheet with information about the researcher journal papers"""
    journal_papers = list(filter(scope_years_paper_or_support, researcher.journal_papers))

    worksheet = workbook.active
    worksheet.title = "Publicações em periódico"

    worksheet.cell(row=1, column=1, value="DOI")
    worksheet.cell(row=1, column=2, value="Autores")
    worksheet.cell(row=1, column=3, value="Título")
    worksheet.cell(row=1, column=4, value="Páginas")
    worksheet.cell(row=1, column=5, value="Ano")
    worksheet.cell(row=1, column=6, value="ISSN")
    worksheet.cell(row=1, column=7, value="Periódico")
    worksheet.cell(row=1, column=8, value="Match")
    worksheet.cell(row=1, column=9, value="JCR")
    worksheet.cell(row=1, column=10, value="Qualis")
    worksheet.cell(row=1, column=11, value="Pontos")

    paper_index = 0
    for row in range(2, len(journal_papers) + 2):
        venue = journal_papers[paper_index].venue

        worksheet.cell(row=row, column=1, value=f'=HYPERLINK("https://www.doi.org/{journal_papers[paper_index].doi}")' if journal_papers[paper_index].doi else '')
        worksheet.cell(row=row, column=2, value=journal_papers[paper_index].authors)
        worksheet.cell(row=row, column=3, value=journal_papers[paper_index].title)
        worksheet.cell(row=row, column=4, value=calculate_number_of_pages(journal_papers[paper_index]))
        worksheet.cell(row=row, column=5, value=journal_papers[paper_index].year)
        worksheet.cell(row=row, column=6, value=venue.issn)
        worksheet.cell(row=row, column=7, value=venue.name)
        worksheet.cell(row=row, column=8, value=venue.official_forum)
        worksheet.cell(row=row, column=9, value=venue.jcr)
        worksheet.cell(row=row, column=10, value=venue.qualis.value)
        worksheet.cell(row=row, column=11, value=get_qualis_points(True, venue.qualis))
        paper_index += 1


def write_conference_papers(researcher, session, workbook):
    """Writes a sheet with information about the researcher conference papers"""
    conference_papers = list(filter(scope_years_paper_or_support, researcher.conference_papers))

    worksheet = workbook.create_sheet("Publicações em conferência")

    worksheet.cell(row=1, column=1, value="DOI")
    worksheet.cell(row=1, column=2, value="Autores")
    worksheet.cell(row=1, column=3, value="Título")
    worksheet.cell(row=1, column=4, value="Páginas")
    worksheet.cell(row=1, column=5, value="Ano")
    worksheet.cell(row=1, column=6, value="Tipo")
    worksheet.cell(row=1, column=7, value="Conferência")
    worksheet.cell(row=1, column=8, value="Match")
    worksheet.cell(row=1, column=9, value="Qualis")
    worksheet.cell(row=1, column=10, value="Pontos")

    paper_index = 0
    for row in range(2, len(conference_papers) + 2):
        venue = conference_papers[paper_index].venue

        worksheet.cell(row=row, column=1, value=f'=HYPERLINK("https://www.doi.org/{conference_papers[paper_index].doi}")' if conference_papers[paper_index].doi else '')
        worksheet.cell(row=row, column=2, value=conference_papers[paper_index].authors)
        worksheet.cell(row=row, column=3, value=conference_papers[paper_index].title)
        worksheet.cell(row=row, column=4, value=calculate_number_of_pages(conference_papers[paper_index]))
        worksheet.cell(row=row, column=5, value=conference_papers[paper_index].year)
        worksheet.cell(row=row, column=6, value=conference_papers[paper_index].nature.value)
        worksheet.cell(row=row, column=7, value=venue.name)
        worksheet.cell(row=row, column=8, value=venue.official_forum)
        worksheet.cell(row=row, column=9, value=venue.qualis.value)
        worksheet.cell(row=row, column=10, value=get_qualis_points(False, venue.qualis))        
        paper_index += 1


def write_researcher_advisements(researcher, session, workbook):
    """Writes a sheet with information about the researcher advisements"""
    researcher_advisements = list(filter(scope_years_paper_or_support, session.query(Advisement).filter(
        Advisement.researcher_id == researcher.id).all()))

    worksheet = workbook.create_sheet("Orientações")

    worksheet.cell(row=1, column=1, value="Tipo")
    worksheet.cell(row=1, column=2, value="Nome do aluno")
    worksheet.cell(row=1, column=3, value="Ano")

    advisement_index = 0
    for row in range(2, len(researcher_advisements) + 2):
        advisement = researcher_advisements[advisement_index]
        worksheet.cell(row=row, column=1, value=advisement.type.value if advisement.type is not None else " ")
        worksheet.cell(row=row, column=2, value=advisement.student_name)
        worksheet.cell(row=row, column=3, value=advisement.year)
        advisement_index += 1


def write_researcher_committee(researcher, session, workbook):
    """Writes a sheet with information about the researcher participation in committees"""
    researcher_committee = list(filter(scope_years_paper_or_support, session.query(Committee).filter(
        Committee.researcher_id == researcher.id).all()))

    worksheet = workbook.create_sheet("Participações em banca")

    worksheet.cell(row=1, column=1, value="Tipo")
    worksheet.cell(row=1, column=2, value="Ano")
    worksheet.cell(row=1, column=3, value="Nome da universidade")
    worksheet.cell(row=1, column=4, value="Nome do aluno ou cargo")

    committee_index = 0
    for row in range(2, len(researcher_committee) + 2):
        worksheet.cell(row=row, column=1, value=researcher_committee[committee_index].type.value)
        worksheet.cell(row=row, column=2, value=researcher_committee[committee_index].year)
        worksheet.cell(row=row, column=3, value=researcher_committee[committee_index].college)
        name_or_position = researcher_committee[committee_index].student_name \
            if researcher_committee[committee_index].type is not CommitteeTypes.CIVIL_SERVICE_EXAMINATION \
            else researcher_committee[committee_index].title
        worksheet.cell(row=row, column=4, value=name_or_position)
        committee_index += 1


def write_researcher_xlsx(researcher, session, workbook):
    """Calls functions to write a .xlsx with information about a researcher"""
    write_journal_papers(researcher, session, workbook)
    write_conference_papers(researcher, session, workbook)
    write_researcher_advisements(researcher, session, workbook)
    write_researcher_committee(researcher, session, workbook)


def write_xlsx_files(researchers_to_write, session):
    """Writes and saves .xlsx files from a researcher list"""
    print("\nGenerating", len(researchers_to_write), "file(s)\n")
    for i in range(len(researchers_to_write)):
        wb = openpyxl.Workbook()
        write_researcher_xlsx(researchers_to_write[i], session, wb)
        wb.save(generate_reseacher_progression_report_output_dir + os.sep + researchers_to_write[i].name + ".xlsx")
    print("Finished generating the file(s)")


def researchers_selection(researchers, args):
    """Makes the user select a researcher to write the .xlsx file from. Or all of them"""
    researcher_index = args

    if researcher_index is None:
        print("Type the number of the researcher you wish to generate the .xlsx file for")
        i = 1
        for i in range(1, len(researchers)+1):
            print(i, " - ", researchers[i-1].name)
        print(i+1, " - All of them")

        researcher_index = input()

        while (not researcher_index.isdigit()) or (int(researcher_index) > len(researchers)+1):
            researcher_index = input("Type a number between 0 and {}".format(i+1))

        researcher_index = int(researcher_index) - 1

    elif isinstance(researcher_index, list): return list(filter(lambda x: researchers_from_args(x, researcher_index), researchers))

    if researcher_index >= len(researchers): return researchers

    return [researchers[researcher_index]]


def print_researchers_ids():
    """Prints the ids which the researchers are going to have on the console"""
    df = pd.read_excel(researchers_file, dtype={'ID Lattes': object})
    for i, row in df.iterrows():
        researcher = row.to_dict()
        if not pd.isnull(researcher['ID Lattes']):
            print("Name:", researcher['Nome'], "/ ID Lattes:", researcher['ID Lattes'], "/ ID database:", str(i+1))
    exit()


def main():
    parser = argparse.ArgumentParser(description="Researchers to generate .xlsx file")
    parser.add_argument("-r", "--researchers", nargs='+')
    parser.add_argument("-a", "--all", action="store_true")
    parser.add_argument("-i", "--ids", action="store_true")

    if parser.parse_args().ids: print_researchers_ids()

    researchers_arg = parser.parse_args().researchers
    if researchers_arg is not None:
        for i in range(len(researchers_arg)):
            if researchers_arg[i].isdigit(): researchers_arg[i] = int(researchers_arg[i])

    session = populate_database.main()
    researchers = session.query(Researcher).all()

    args = len(researchers) if parser.parse_args().all else researchers_arg
    researchers_to_write = researchers_selection(researchers, args)
    write_xlsx_files(researchers_to_write, session)


if __name__ == "__main__":
    main()