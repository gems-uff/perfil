import openpyxl
import os
import populate
from sqlalchemy import and_
from config import full_professorship_advancement_output_dir, start_year, end_year
from database.database_manager import Researcher, Journal, Conference, ResearcherAdvisement, ResearcherCommittee
from database.titles_support import CommitteeTypes


def correct_year(paper_or_support):
    """filter function to get only database objects within the years specified"""
    if start_year <= paper_or_support.year <= end_year: return True
    return False


def calculate_number_of_pages(paper):
    """Calculates the number of pages of a given paper"""
    if (paper.last_page is None) or (paper.first_page is None) \
            or (type(paper.last_page) is not int) or (type(paper.first_page) is not int): return ""
    return paper.last_page - paper.first_page + 1


def write_journal_papers(researcher, session, workbook):
    """Writes a sheet with information about the researcher journal papers"""
    journal_papers = list(filter(correct_year, researcher.journal_papers))

    worksheet = workbook.active
    worksheet.title = "Publicações em periódico"

    worksheet.cell(row=1, column=1, value="Título")
    worksheet.cell(row=1, column=2, value="Autores")
    worksheet.cell(row=1, column=3, value="Periódico")
    worksheet.cell(row=1, column=4, value="Ano")
    worksheet.cell(row=1, column=5, value="Qualis")
    worksheet.cell(row=1, column=6, value="Número de páginas")
    worksheet.cell(row=1, column=7, value="JCR")
    worksheet.cell(row=1, column=8, value="ISSN")

    paper_index = 0
    for row in range(2, len(journal_papers) + 2):
        venue = session.query(Journal).filter(Journal.id == journal_papers[paper_index].venue).all()

        worksheet.cell(row=row, column=1, value=journal_papers[paper_index].title)
        worksheet.cell(row=row, column=2, value=journal_papers[paper_index].authors)
        worksheet.cell(row=row, column=3, value=venue[0].name)
        worksheet.cell(row=row, column=4, value=journal_papers[paper_index].year)
        qualis = venue[0].qualis.value if venue[0].qualis is not None else ""
        worksheet.cell(row=row, column=5, value=qualis)
        worksheet.cell(row=row, column=6, value=calculate_number_of_pages(journal_papers[paper_index]))
        worksheet.cell(row=row, column=7, value=venue[0].jcr)
        worksheet.cell(row=row, column=8, value=venue[0].issn)
        paper_index += 1


def write_conference_papers(researcher, session, workbook):
    """Writes a sheet with information about the researcher conference papers"""
    conference_papers = list(filter(correct_year, researcher.conference_papers))

    worksheet = workbook.create_sheet("Publicações em conferência")

    worksheet.cell(row=1, column=1, value="Título")
    worksheet.cell(row=1, column=2, value="Autores")
    worksheet.cell(row=1, column=3, value="Conferência")
    worksheet.cell(row=1, column=4, value="Ano")
    worksheet.cell(row=1, column=5, value="Qualis")
    worksheet.cell(row=1, column=6, value="Número de páginas")
    worksheet.cell(row=1, column=7, value="Tipo")

    paper_index = 0
    for row in range(2, len(conference_papers) + 2):
        venue = session.query(Conference).filter(Conference.id == conference_papers[paper_index].venue).all()

        worksheet.cell(row=row, column=1, value=conference_papers[paper_index].title)
        worksheet.cell(row=row, column=2, value=conference_papers[paper_index].authors)
        worksheet.cell(row=row, column=3, value=venue[0].name)
        worksheet.cell(row=row, column=4, value=conference_papers[paper_index].year)
        qualis = venue[0].qualis.value if venue[0].qualis is not None else ""
        worksheet.cell(row=row, column=5, value=qualis)
        worksheet.cell(row=row, column=6, value=calculate_number_of_pages(conference_papers[paper_index]))
        worksheet.cell(row=row, column=7, value=conference_papers[paper_index].nature.value)
        paper_index += 1


def write_researcher_advisements(researcher, session, workbook):
    """Writes a sheet with information about the researcher advisements"""
    researcher_advisements = list(filter(correct_year, session.query(ResearcherAdvisement).filter(
        ResearcherAdvisement.researcher_id == researcher.id).all()))

    worksheet = workbook.create_sheet("Orientações")

    worksheet.cell(row=1, column=1, value="Tipo")
    worksheet.cell(row=1, column=2, value="Nome do aluno")
    worksheet.cell(row=1, column=3, value="Ano")

    advisement_index = 0
    for row in range(2, len(researcher_advisements) + 2):
        worksheet.cell(row=row, column=1, value=researcher_advisements[advisement_index].type.value)
        worksheet.cell(row=row, column=2, value=researcher_advisements[advisement_index].student_name)
        worksheet.cell(row=row, column=3, value=researcher_advisements[advisement_index].year)
        advisement_index += 1


def write_researcher_committee(researcher, session, workbook):
    """Writes a sheet with information about the researcher participation in committees"""
    researcher_committee = list(filter(correct_year, session.query(ResearcherCommittee).filter(
        ResearcherCommittee.researcher_id == researcher.id).all()))

    worksheet = workbook.create_sheet("Participações em banca")

    worksheet.cell(row=1, column=1, value="Tipo")
    worksheet.cell(row=1, column=2, value="Ano")
    worksheet.cell(row=1, column=3, value="Nome da universidade")
    worksheet.cell(row=1, column=4, value="Nome do aluno ou cargo")

    committee_index = 0
    for row in range(2, len(researcher_committee) + 2):
        worksheet.cell(row=row, column=1, value=researcher_committee[committee_index].type.value)
        worksheet.cell(row=row, column=2, value=researcher_committee[committee_index].year)
        worksheet.cell(row=row, column=3, value=researcher_committee[committee_index].college)
        name_or_position = researcher_committee[committee_index].student_name \
            if researcher_committee[committee_index].type is not CommitteeTypes.CIVIL_SERVICE_EXAMINATION \
            else researcher_committee[committee_index].title
        worksheet.cell(row=row, column=4, value=name_or_position)
        committee_index += 1


def write_researcher_xlsx(researcher, session, workbook):
    """Calls functions to write a .xlsx with information about a researcher"""
    write_journal_papers(researcher, session, workbook)
    write_conference_papers(researcher, session, workbook)
    write_researcher_advisements(researcher, session, workbook)
    write_researcher_committee(researcher, session, workbook)


def write_xlsx_files(researchers_to_write, session):
    """Writes and saves .xlsx files from a researcher list"""
    for i in range(len(researchers_to_write)):
        wb = openpyxl.Workbook()
        write_researcher_xlsx(researchers_to_write[i], session, wb)
        wb.save(full_professorship_advancement_output_dir + os.sep + researchers_to_write[i].name+".xlsx")


def researchers_selection(researchers):
    """Makes the user select a researcher to write the .xlsx file from. Or all of them"""
    print("Type the number of the researcher you wish to generate the .xlsx file for")
    i = 0
    for i in range(len(researchers)):
        print(i, " - ", researchers[i].name)
    print(i+1, " - All of them")

    researcher_index = input()

    while (not researcher_index.isdigit()) or (int(researcher_index) > len(researchers)):
        researcher_index = input("Type a number between 0 and {}".format(i+1))

    if int(researcher_index) >= len(researchers): return researchers

    return [researchers[int(researcher_index)]]


def main():
    session = populate.main()
    researchers = session.query(Researcher).all()

    researchers_to_write = researchers_selection(researchers)
    write_xlsx_files(researchers_to_write, session)


if __name__ == "__main__":
    main()